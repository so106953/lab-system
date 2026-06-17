import streamlit as st
from supabase import create_client, Client
import pandas as pd
import base64
import io
from datetime import datetime, timedelta
import pytz
import time
from openpyxl.styles import PatternFill

# --- 1. 数据库配置 ---
URL = "https://xcrbvvlsbjsmxaepkdbl.supabase.co"
KEY = "sb_publishable_QYXt0Fs5YKpCBXxCjdb4sg_9y7rkksj"
supabase: Client = create_client(URL, KEY)

# --- 2. 组别与设备清单配置 ---
# 领用人所属组别清单
ALL_GROUPS = ["components", "PV", "tools", "HVAC", "Lingting", "PC组", "小家电", "洗衣机", "材料", "电缆", "电池组", "hiteck组"]

# 各组资产所属清单
GROUP_CONFIG = {
    "components": ["万用表", "示波器", "电流钳", "环境箱", "小烤箱", "游标卡尺", "直流源", "电子负载", "datalogger", "电压探头", "高压探头"],
    "电池组": ["万用表", "内阻仪", "充放电柜", "示波器", "环境箱"],
    "hiteck组": ["示波器", "万用表", "电子负载", "直流电源"],
    "PV": ["光伏测试仪", "万用表", "直流电源", "电子负载"],
    "tools": ["游标卡尺", "扭力计", "示波器", "万用表"],
    "HVAC": ["环境箱", "温度记录仪", "压力计", "流量计"],
    "Lingting": ["功率计", "积分球", "光谱仪", "万用表"],
    "PC组": ["示波器", "万用表", "电子负载", "Datalogger"],
    "小家电": ["功率计", "安规测试仪", "温升仪", "示波器"],
    "洗衣机": ["环境箱", "流量计", "转速表", "温升仪"],
    "材料": ["拉力计", "硬度计", "烤箱", "游标卡尺"],
    "电缆": ["弯曲试验机", "耐压仪", "电阻测试仪", "卡尺"],
    "默认": ["直流电源", "万用表", "示波器"]
}

# --- 3. 极致视觉定制 ---
st.set_page_config(page_title="UL设备登记系统", layout="centered")

def get_base64_image(file_path):
    try:
        with open(file_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except: return ""

img_base64 = get_base64_image("IMG_4614.jpeg")

style = f"""
    <style>
    header, footer, [data-testid="stHeader"], [data-testid="stToolbar"] {{display: none !important;}}
    [data-testid="stStatusWidget"], [data-testid="manage-app-button"], .viewerBadge_container__1QSob {{
        display: none !important; position: fixed !important; left: -9999px !important;
    }}
    :root {{ --bg-color: #FFFFFF; --text-color: #000000; --card-bg: #FFFFFF; }}
    @media (prefers-color-scheme: dark) {{
        :root {{ --bg-color: #0E1117; --text-color: #FFFFFF; --card-bg: #1A1C23; }}
        input, .stSelectbox div {{ background-color: #2D2F39 !important; color: white !important; }}
    }}
    html, body, [data-testid="stAppViewContainer"], .stApp {{ background-color: var(--bg-color) !important; color: var(--text-color) !important; }}
    h1, h2, label, span, div {{ color: var(--text-color) !important; font-weight: 600 !important; }}
    .logo-box {{ display: flex; justify-content: center; padding: 15px 0; }}
    .logo-box img {{ max-width: 180px; height: auto; }}
    .stApp::before {{
        content: ""; position: fixed; top: 0; left: 0; bottom: 0; right: 0;
        background-image: url("data:image/jpeg;base64,{img_base64}");
        background-repeat: no-repeat; background-position: center;
        background-size: 55%; opacity: 0.02; z-index: -1;
    }}
    div[data-testid="stForm"] {{ border-radius: 24px !important; background-color: var(--card-bg) !important; padding: 25px !important; box-shadow: 0 10px 40px rgba(0,0,0,0.1) !important; border: 1px solid rgba(128,128,128,0.1) !important; }}
    .stButton>button {{ width: 100%; border-radius: 12px !important; height: 3.8em !important; background-color: #28A745 !important; color: #FFFFFF !important; font-weight: bold !important; }}
    </style>
"""
st.markdown(style, unsafe_allow_html=True)

# --- 4. 辅助格式化函数 ---
def format_time_diff(td):
    total_seconds = int(abs(td.total_seconds()))
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, _ = divmod(remainder, 60)
    res = []
    if days > 0: res.append(f"{days}天")
    if hours > 0: res.append(f"{hours}小时")
    if minutes > 0: res.append(f"{minutes}分")
    return "".join(res) if res else "刚刚"

def format_loan_duration(d):
    if not d or d == 0: return "-"
    total_mins = int(round(d * 1440))
    if total_mins < 60: return f"{total_mins}分钟"
    return f"{round(d, 2)}天"

# --- 5. 核心逻辑：自动超时分级提醒 ---
def monitor_loan_status(owning_group):
    try:
        res = supabase.table("lab_records").select("*").execute()
        if res.data:
            df = pd.DataFrame(res.data)
            latest = df.sort_values('created_at').groupby('device_id').last().reset_index()
            borrowed = latest[latest['action_type'] == '领用']
            if owning_group != "默认":
                borrowed = borrowed[borrowed['lab_group'] == owning_group]
            
            now = datetime.now(pytz.timezone('Asia/Shanghai'))
            red_list, yellow_list = [], []

            for _, row in borrowed.iterrows():
                start_time = pd.to_datetime(row['created_at']).astimezone(pytz.timezone('Asia/Shanghai'))
                limit_days = float(row['loan_days']) if row['loan_days'] else 1.0
                due_time = start_time + timedelta(days=limit_days)
                
                person_info = f"**{row['staff_id']}** ({row['staff_group']})"
                
                if now >= due_time:
                    overdue_str = format_time_diff(now - due_time)
                    red_list.append(f"🔴 **{row['device_name']}** ({row['device_id']}) | 负责人: {person_info} | 已超期: **{overdue_str}**")
                elif now >= (due_time - timedelta(minutes=5)):
                    remain_str = format_time_diff(due_time - now)
                    yellow_list.append(f"🟡 **{row['device_name']}** ({row['device_id']}) | 负责人: {person_info} | 剩余: **{remain_str}**")

            if red_list:
                st.error("🚨 **设备超时未还警示名单 (Urgent Overdue)**")
                for msg in red_list: st.markdown(f"<div style='padding:8px; border-left:5px solid red; background:rgba(255,0,0,0.05); margin-bottom:5px; font-size:14px;'>{msg}</div>", unsafe_allow_html=True)
            if yellow_list:
                st.warning("⏳ **即将超时提醒 (Due Soon)**")
                for msg in yellow_list: st.markdown(f"<div style='padding:8px; border-left:5px solid #FFCC00; background:rgba(255,204,0,0.05); margin-bottom:5px; font-size:14px;'>{msg}</div>", unsafe_allow_html=True)
    except: pass

# --- 6. 页面入口 ---
query_params = st.query_params
owning_group = query_params.get("group", "默认")
monitor_loan_status(owning_group)

if img_base64:
    st.markdown(f'<div class="logo-box"><img src="data:image/jpeg;base64,{img_base64}"></div>', unsafe_allow_html=True)
st.markdown(f"<h2 style='text-align: center; margin-top: 0px;'>{owning_group} 资产登记表</h2>", unsafe_allow_html=True)

with st.form("lab_form", clear_on_submit=True):
    staff_id = st.text_input("👤 领用人工号 (Staff ID)", placeholder="请输入工号")
    staff_group = st.selectbox("🏢 领用人所属组别", ["请选择组别"] + ALL_GROUPS)
    st.markdown("**⏳ 预计借用时长 (Loan Duration)**")
    c1, c2 = st.columns([2, 1])
    with c2: unit = st.radio("单位", ["天", "分钟"], horizontal=True)
    with c1:
        if unit == "分钟":
            amt = st.number_input("数值", min_value=10, value=30, step=10)
            final_days = amt / 1440.0
        else:
            amt = st.number_input("数值", min_value=0.1, value=1.0, step=0.5)
            final_days = float(amt)

    st.markdown("---")
    action_type = st.radio("📝 操作类型", ["领用 (Check-out)", "归还 (Return)"], horizontal=True)
    available_devices = GROUP_CONFIG.get(owning_group, GROUP_CONFIG["默认"])
    device_name = st.selectbox("📦 资产名称", ["请选择资产"] + available_devices)
    device_id = st.text_input("🔢 资产编号 (SN)", placeholder="输入唯一编号")
    submit_btn = st.form_submit_button("确认提交登记 (SUBMIT)")

if submit_btn:
    if not staff_id or staff_group == "请选择组别" or device_name == "请选择资产" or not device_id:
        st.error("❌ 请完整填写所有必填项！")
    else:
        try:
            entry = {
                "staff_id": staff_id, "staff_group": staff_group, "lab_group": owning_group,
                "action_type": "领用" if "领用" in action_type else "归还",
                "device_name": device_name, "device_id": device_id,
                "loan_days": final_days if "领用" in action_type else 0
            }
            supabase.table("lab_records").insert(entry).execute()
            st.balloons()
            st.success("✅ 登记成功！数据已实时备份。")
            time.sleep(2)
            st.rerun()
        except Exception as e: st.error(f"失败: {e}")

# --- 7. 管理员后台 ---
st.markdown("<br>", unsafe_allow_html=True)
with st.expander("📊 查看记录 (带色彩 Excel 导出)"):
    try:
        res = supabase.table("lab_records").select("*").order("created_at", desc=True).execute()
        if res.data:
            df = pd.DataFrame(res.data)
            df['created_at'] = pd.to_datetime(df['created_at']).dt.tz_convert('Asia/Shanghai').dt.strftime('%Y-%m-%d %H:%M:%S')
            if owning_group != "默认":
                df = df[df['lab_group'] == owning_group]
            
            df['借用时长'] = df['loan_days'].apply(format_loan_duration)
            df_display = df[['staff_id', 'staff_group', 'action_type', 'device_name', 'device_id', '借用时长', 'created_at']]
            df_display.columns = ["工号", "人员组别", "类型", "资产名称", "资产编号", "借用时长", "登记时间"]
            st.dataframe(df_display, use_container_width=True)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_display.to_excel(writer, index=False, sheet_name='Records')
                ws = writer.sheets['Records']
                yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                for r_idx, action in enumerate(df_display["类型"], start=2):
                    fill = yellow if action == "领用" else green
                    for cell in ws[r_idx]: cell.fill = fill
            st.download_button(f"📥 导出 {owning_group} Excel", output.getvalue(), f"UL_{owning_group}_Records.xlsx")
    except: st.write("暂无记录。")